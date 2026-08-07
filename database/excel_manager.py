"""
excel_manager.py

Handles all Excel database operations.

Ramdev Billing Software
"""

import os
from openpyxl import Workbook, load_workbook

from .schema import (
    COMPANY_SHEET,
    DATABASE_NAME,
    DEFAULT_COMPANY,
    DEFAULT_SETTINGS,
    DEFAULT_USER,
    SETTINGS_SHEET,
    USERS_SHEET,
    WORKBOOK_STRUCTURE
)


class ExcelManager:
    """
    Generic Excel Database Manager
    """

    def __init__(self):

        self.workbook = None

        self.load_database()

    # =====================================================
    # DATABASE
    # =====================================================

    def load_database(self):
        """
        Load existing Excel database.

        If Database.xlsx does not exist,
        create a new database.

        Also creates any missing sheets
        defined in schema.py.
        """

        if os.path.exists(DATABASE_NAME):

            self.workbook = load_workbook(
                DATABASE_NAME
            )

        else:

            self.workbook = Workbook()

            # Remove default Sheet
            default_sheet = self.workbook.active

            self.workbook.remove(
                default_sheet
            )

        # Create missing sheets
        self.create_missing_sheets()

        # Populate defaults for empty sheets
        self.populate_default_data()

        # Save database
        self.save_database()

    # =====================================================

    def create_missing_sheets(self):
        """
        Create all sheets defined in schema.py
        if they do not already exist.
        """

        for sheet_name, headers in WORKBOOK_STRUCTURE.items():

            if sheet_name not in self.workbook.sheetnames:

                sheet = self.workbook.create_sheet(
                    sheet_name
                )

                # Add headers
                for column, header in enumerate(
                    headers,
                    start=1
                ):

                    sheet.cell(
                        row=1,
                        column=column,
                        value=header
                    )

    # =====================================================

    def save_database(self):
        """
        Save workbook to Database.xlsx.
        """

        self.workbook.save(
            DATABASE_NAME
        )

    # =====================================================
    # SHEET
    # =====================================================

    def get_sheet(self, sheet_name):
        """
        Return worksheet.
        """

        if sheet_name not in self.workbook.sheetnames:

            if sheet_name in WORKBOOK_STRUCTURE:

                headers = WORKBOOK_STRUCTURE[
                    sheet_name
                ]

                sheet = self.workbook.create_sheet(
                    sheet_name
                )

                for column, header in enumerate(
                    headers,
                    start=1
                ):

                    sheet.cell(
                        row=1,
                        column=column,
                        value=header
                    )

                self.save_database()

            else:

                raise ValueError(
                    f"Sheet '{sheet_name}' does not exist."
                )

        return self.workbook[sheet_name]

    # =====================================================

    def populate_default_data(self):
        """
        Populate default demo data for new or empty sheets.
        """

        company_sheet = self.get_sheet(COMPANY_SHEET)
        users_sheet = self.get_sheet(USERS_SHEET)
        settings_sheet = self.get_sheet(SETTINGS_SHEET)

        if company_sheet.max_row == 1:
            for column, value in enumerate(
                DEFAULT_COMPANY,
                start=1
            ):
                company_sheet.cell(
                    row=2,
                    column=column,
                    value=value
                )

        if users_sheet.max_row == 1:
            for column, value in enumerate(
                DEFAULT_USER,
                start=1
            ):
                users_sheet.cell(
                    row=2,
                    column=column,
                    value=value
                )

        if settings_sheet.max_row == 1:
            row = 2
            for key, value in DEFAULT_SETTINGS:
                settings_sheet.cell(row=row, column=1).value = key
                settings_sheet.cell(row=row, column=2).value = value
                row += 1

    def get_headers(self, sheet_name):
        """
        Return header row.
        """

        sheet = self.get_sheet(
            sheet_name
        )

        return [
            cell.value
            for cell in sheet[1]
        ]

    # =====================================================
    # READ
    # =====================================================

    def get_all_records(self, sheet_name):
        """
        Return all records as a list
        of dictionaries.
        """

        sheet = self.get_sheet(
            sheet_name
        )

        headers = self.get_headers(
            sheet_name
        )

        records = []

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            # Skip completely empty rows
            if all(
                value is None
                for value in row
            ):
                continue

            record = dict(
                zip(
                    headers,
                    row
                )
            )

            records.append(
                record
            )

        return records

    # =====================================================
    # INSERT
    # =====================================================

    def insert_record(
        self,
        sheet_name,
        data
    ):
        """
        Insert a new record.

        Example:

        data = {
            "Customer ID": "C0001",
            "Customer Name": "ABC"
        }
        """

        sheet = self.get_sheet(
            sheet_name
        )

        headers = self.get_headers(
            sheet_name
        )

        row = []

        for header in headers:

            row.append(
                data.get(
                    header,
                    ""
                )
            )

        sheet.append(row)

        self.save_database()

        return True

    # =====================================================
    # UPDATE
    # =====================================================

    def update_record(
        self,
        sheet_name,
        key_column,
        key_value,
        data
    ):
        """
        Update an existing record.
        """

        sheet = self.get_sheet(
            sheet_name
        )

        headers = self.get_headers(
            sheet_name
        )

        if key_column not in headers:

            raise ValueError(
                f"Column '{key_column}' "
                f"does not exist in '{sheet_name}'."
            )

        key_index = (
            headers.index(key_column) + 1
        )

        for row in range(
            2,
            sheet.max_row + 1
        ):

            current_value = sheet.cell(
                row=row,
                column=key_index
            ).value

            if str(current_value) == str(
                key_value
            ):

                for column, header in enumerate(
                    headers,
                    start=1
                ):

                    if header in data:

                        sheet.cell(
                            row=row,
                            column=column
                        ).value = data[header]

                self.save_database()

                return True

        return False

    # =====================================================
    # DELETE
    # =====================================================

    def delete_record(
        self,
        sheet_name,
        key_column,
        key_value
    ):
        """
        Delete an existing record.
        """

        sheet = self.get_sheet(
            sheet_name
        )

        headers = self.get_headers(
            sheet_name
        )

        if key_column not in headers:

            raise ValueError(
                f"Column '{key_column}' "
                f"does not exist in '{sheet_name}'."
            )

        key_index = (
            headers.index(key_column) + 1
        )

        for row in range(
            2,
            sheet.max_row + 1
        ):

            current_value = sheet.cell(
                row=row,
                column=key_index
            ).value

            if str(current_value) == str(
                key_value
            ):

                sheet.delete_rows(row)

                self.save_database()

                return True

        return False

    # =====================================================
    # FIND
    # =====================================================

    def find_record(
        self,
        sheet_name,
        key_column,
        key_value
    ):
        """
        Find one record.
        """

        sheet = self.get_sheet(
            sheet_name
        )

        headers = self.get_headers(
            sheet_name
        )

        if key_column not in headers:

            raise ValueError(
                f"Column '{key_column}' "
                f"does not exist in '{sheet_name}'."
            )

        key_index = (
            headers.index(key_column) + 1
        )

        for row in range(
            2,
            sheet.max_row + 1
        ):

            current_value = sheet.cell(
                row=row,
                column=key_index
            ).value

            if str(current_value) == str(
                key_value
            ):

                values = []

                for column in range(
                    1,
                    len(headers) + 1
                ):

                    values.append(
                        sheet.cell(
                            row=row,
                            column=column
                        ).value
                    )

                return dict(
                    zip(
                        headers,
                        values
                    )
                )

        return None

    # =====================================================
    # EXISTS
    # =====================================================

    def record_exists(
        self,
        sheet_name,
        key_column,
        key_value
    ):
        """
        Check whether a record exists.
        """

        return (
            self.find_record(
                sheet_name,
                key_column,
                key_value
            )
            is not None
        )

    # =====================================================
    # GENERATE ID
    # =====================================================

    def generate_next_id(
        self,
        sheet_name,
        prefix,
        id_column
    ):
        """
        Generate next sequential ID.

        Examples:

        CUST0001
        PROD0001
        PUR0001
        INV0001
        """

        records = self.get_all_records(
            sheet_name
        )

        if not records:

            return f"{prefix}0001"

        last_number = 0

        for record in records:

            value = str(
                record.get(
                    id_column,
                    ""
                )
            )

            if value.startswith(prefix):

                try:

                    number = int(
                        value[
                            len(prefix):
                        ]
                    )

                    if number > last_number:

                        last_number = number

                except ValueError:

                    continue

        return (
            f"{prefix}"
            f"{last_number + 1:04d}"
        )