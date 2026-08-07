"""
create_database.py

Creates the Excel database automatically with all required sheets,
headers, and default data.

Ramdev Billing Software
"""

import os
from openpyxl import Workbook

from .schema import (
    DATABASE_FOLDER,
    DATABASE_NAME,
    WORKBOOK_STRUCTURE,

    COMPANY_SHEET,
    USERS_SHEET,
    SETTINGS_SHEET,

    DEFAULT_COMPANY,
    DEFAULT_USER,
    DEFAULT_SETTINGS
)


class DatabaseCreator:

    def __init__(self):
        os.makedirs(DATABASE_FOLDER, exist_ok=True)

    def create_database(self):
        """
        Create Database.xlsx if it does not exist.
        """

        if os.path.exists(DATABASE_NAME):
            return False

        workbook = Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)

        # Create all sheets
        for sheet_name, headers in WORKBOOK_STRUCTURE.items():

            sheet = workbook.create_sheet(title=sheet_name)

            # Write headers
            for column, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=column).value = header

            # Insert default data
            if sheet_name == COMPANY_SHEET:
                for column, value in enumerate(DEFAULT_COMPANY, start=1):
                    sheet.cell(row=2, column=column).value = value

            elif sheet_name == USERS_SHEET:
                for column, value in enumerate(DEFAULT_USER, start=1):
                    sheet.cell(row=2, column=column).value = value

            elif sheet_name == SETTINGS_SHEET:
                row = 2
                for setting in DEFAULT_SETTINGS:
                    sheet.cell(row=row, column=1).value = setting[0]
                    sheet.cell(row=row, column=2).value = setting[1]
                    row += 1

        workbook.save(DATABASE_NAME)

        return True


def initialize_database():
    """
    Initialize database.
    """

    creator = DatabaseCreator()
    creator.create_database()


if __name__ == "__main__":
    initialize_database()