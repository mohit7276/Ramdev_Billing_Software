"""
excel_invoice_export.py

Generate formatted Excel Invoice

Ramdev Billing Software
"""

import os

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)


class ExcelInvoiceExporter:

    def __init__(self):

        self.header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        self.header_font = Font(
            bold=True,
            color="FFFFFF",
            size=12
        )

        self.bold_font = Font(
            bold=True
        )

        self.title_font = Font(
            bold=True,
            size=18
        )

        thin = Side(style="thin")

        self.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

    # --------------------------------------------------------

    def generate(
        self,
        company,
        customer,
        invoice,
        items,
        output_file
    ):

        os.makedirs(
            os.path.dirname(output_file),
            exist_ok=True
        )

        wb = Workbook()

        ws = wb.active

        ws.title = "Invoice"

        # ============================================
        # Company Details
        # ============================================

        ws["A1"] = company["Company Name"]
        ws["A1"].font = self.title_font

        ws["A2"] = company["Address"]
        ws["A3"] = f"Phone : {company['Phone']}"
        ws["A4"] = f"GSTIN : {company['GSTIN']}"

        # ============================================
        # Invoice Details
        # ============================================

        ws["F1"] = "Invoice No"
        ws["G1"] = invoice["Invoice Number"]

        ws["F2"] = "Date"
        ws["G2"] = invoice["Invoice Date"]

        ws["F3"] = "Customer"
        ws["G3"] = customer["Customer Name"]

        ws["F4"] = "Phone"
        ws["G4"] = customer["Phone Number"]

        # ============================================
        # Table Header
        # ============================================

        headers = [

            "Product",

            "Color",

            "Size",

            "Qty",

            "Rate",

            "GST",

            "Amount"

        ]

        row = 7

        for col, text in enumerate(headers, start=1):

            cell = ws.cell(row=row, column=col)

            cell.value = text

            cell.fill = self.header_fill

            cell.font = self.header_font

            cell.alignment = Alignment(horizontal="center")

            cell.border = self.border

        # ============================================
        # Items
        # ============================================

        row = 8

        for item in items:

            amount = item["Quantity"] * item["Rate"]

            gst = amount * item["GST %"] / 100

            total = amount + gst

            values = [

                item["Product Name"],

                item["Color"],

                item["Size"],

                item["Quantity"],

                item["Rate"],

                item["GST %"],

                total

            ]

            for col, value in enumerate(values, start=1):

                cell = ws.cell(row=row, column=col)

                cell.value = value

                cell.border = self.border

            row += 1

        # ============================================
        # Totals
        # ============================================

        row += 2

        ws.cell(row=row, column=6).value = "Subtotal"
        ws.cell(row=row, column=7).value = invoice["Subtotal"]

        row += 1

        ws.cell(row=row, column=6).value = "GST"
        ws.cell(row=row, column=7).value = invoice["GST"]

        row += 1

        ws.cell(row=row, column=6).value = "Discount"
        ws.cell(row=row, column=7).value = invoice["Discount"]

        row += 1

        ws.cell(row=row, column=6).value = "Grand Total"

        total_cell = ws.cell(row=row, column=7)

        total_cell.value = invoice["Grand Total"]

        total_cell.font = Font(
            bold=True,
            size=13
        )

        # ============================================
        # Auto Width
        # ============================================

        for column in ws.columns:

            length = 0

            letter = column[0].column_letter

            for cell in column:

                try:

                    length = max(
                        length,
                        len(str(cell.value))
                    )

                except:

                    pass

            ws.column_dimensions[letter].width = length + 4

        wb.save(output_file)

        return output_file
