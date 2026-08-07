"""
invoice_export.py

Exports invoices to PDF and Excel.

Ramdev Billing Software
"""

import os

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


class InvoiceExporter:

    # =====================================================
    # EXPORT EXCEL
    # =====================================================

    @staticmethod
    def export_excel(
        invoice,
        items,
        folder="Invoices"
    ):

        os.makedirs(
            folder,
            exist_ok=True
        )

        invoice_number = str(
            invoice["Invoice Number"]
        )

        path = os.path.join(
            folder,
            f"{invoice_number}.xlsx"
        )

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Invoice"

        # -------------------------------------------------
        # Invoice Details
        # -------------------------------------------------

        details = [

            ("Invoice Number",
             invoice.get(
                 "Invoice Number",
                 ""
             )),

            ("Invoice Date",
             invoice.get(
                 "Invoice Date",
                 ""
             )),

            ("Invoice Time",
             invoice.get(
                 "Invoice Time",
                 ""
             )),

            ("Customer Name",
             invoice.get(
                 "Customer Name",
                 ""
             )),

            ("Phone Number",
             invoice.get(
                 "Phone Number",
                 ""
             )),

            ("Payment Mode",
             invoice.get(
                 "Payment Mode",
                 ""
             )),

            ("Payment Status",
             invoice.get(
                 "Payment Status",
                 ""
             ))
        ]

        row = 1

        for key, value in details:

            sheet.cell(
                row=row,
                column=1,
                value=key
            )

            sheet.cell(
                row=row,
                column=2,
                value=value
            )

            row += 1

        row += 1

        # -------------------------------------------------
        # Items Header
        # -------------------------------------------------

        headers = [
            "Product ID",
            "Product Name",
            "Color",
            "Size",
            "Quantity",
            "Rate",
            "GST %",
            "Amount"
        ]

        for column, header in enumerate(
            headers,
            start=1
        ):

            sheet.cell(
                row=row,
                column=column,
                value=header
            )

        row += 1

        # -------------------------------------------------
        # Items
        # -------------------------------------------------

        for item in items:

            values = [

                item.get(
                    "Product ID",
                    ""
                ),

                item.get(
                    "Product Name",
                    ""
                ),

                item.get(
                    "Color",
                    ""
                ),

                item.get(
                    "Size",
                    ""
                ),

                item.get(
                    "Quantity",
                    0
                ),

                item.get(
                    "Rate",
                    0
                ),

                item.get(
                    "GST %",
                    0
                ),

                item.get(
                    "Amount",
                    0
                )
            ]

            for column, value in enumerate(
                values,
                start=1
            ):

                sheet.cell(
                    row=row,
                    column=column,
                    value=value
                )

            row += 1

        row += 1

        # -------------------------------------------------
        # Totals
        # -------------------------------------------------

        totals = [

            (
                "Subtotal",
                invoice.get(
                    "Subtotal",
                    0
                )
            ),

            (
                "Discount",
                invoice.get(
                    "Discount",
                    0
                )
            ),

            (
                "GST",
                invoice.get(
                    "GST",
                    0
                )
            ),

            (
                "Grand Total",
                invoice.get(
                    "Grand Total",
                    0
                )
            )
        ]

        for key, value in totals:

            sheet.cell(
                row=row,
                column=7,
                value=key
            )

            sheet.cell(
                row=row,
                column=8,
                value=value
            )

            row += 1

        # -------------------------------------------------
        # Column Widths
        # -------------------------------------------------

        widths = {
            "A": 18,
            "B": 25,
            "C": 15,
            "D": 15,
            "E": 12,
            "F": 12,
            "G": 12,
            "H": 15
        }

        for column, width in widths.items():

            sheet.column_dimensions[
                column
            ].width = width

        workbook.save(path)

        return path

    # =====================================================
    # EXPORT PDF
    # =====================================================

    @staticmethod
    def export_pdf(
        invoice,
        items,
        folder="Invoices"
    ):

        os.makedirs(
            folder,
            exist_ok=True
        )

        invoice_number = str(
            invoice["Invoice Number"]
        )

        path = os.path.join(
            folder,
            f"{invoice_number}.pdf"
        )

        pdf = canvas.Canvas(
            path,
            pagesize=A4
        )

        width, height = A4

        y = height - 50

        # -------------------------------------------------
        # Title
        # -------------------------------------------------

        pdf.setFont(
            "Helvetica-Bold",
            18
        )

        pdf.drawString(
            50,
            y,
            "RAMDEV ENTERPRISES"
        )

        y -= 30

        pdf.setFont(
            "Helvetica-Bold",
            13
        )

        pdf.drawString(
            50,
            y,
            "INVOICE"
        )

        y -= 30

        # -------------------------------------------------
        # Invoice Details
        # -------------------------------------------------

        pdf.setFont(
            "Helvetica",
            10
        )

        details = [

            f"Invoice No: {invoice.get('Invoice Number', '')}",

            f"Date: {invoice.get('Invoice Date', '')}",

            f"Customer: {invoice.get('Customer Name', '')}",

            f"Phone: {invoice.get('Phone Number', '')}",

            f"Payment: {invoice.get('Payment Mode', '')}"
        ]

        for line in details:

            pdf.drawString(
                50,
                y,
                line
            )

            y -= 18

        y -= 15

        # -------------------------------------------------
        # Table Header
        # -------------------------------------------------

        pdf.setFont(
            "Helvetica-Bold",
            9
        )

        columns = [
            ("Product", 50),
            ("Qty", 280),
            ("Rate", 330),
            ("GST", 390),
            ("Amount", 470)
        ]

        for text, x in columns:

            pdf.drawString(
                x,
                y,
                text
            )

        y -= 18

        pdf.setFont(
            "Helvetica",
            9
        )

        # -------------------------------------------------
        # Items
        # -------------------------------------------------

        for item in items:

            product_name = str(
                item.get(
                    "Product Name",
                    ""
                )
            )

            quantity = str(
                item.get(
                    "Quantity",
                    0
                )
            )

            rate = str(
                item.get(
                    "Rate",
                    0
                )
            )

            gst = str(
                item.get(
                    "GST %",
                    0
                )
            )

            amount = str(
                item.get(
                    "Amount",
                    0
                )
            )

            pdf.drawString(
                50,
                y,
                product_name[:35]
            )

            pdf.drawString(
                280,
                y,
                quantity
            )

            pdf.drawString(
                330,
                y,
                rate
            )

            pdf.drawString(
                390,
                y,
                gst + "%"
            )

            pdf.drawString(
                470,
                y,
                amount
            )

            y -= 18

            if y < 100:

                pdf.showPage()

                y = height - 50

                pdf.setFont(
                    "Helvetica",
                    9
                )

        # -------------------------------------------------
        # Totals
        # -------------------------------------------------

        y -= 15

        pdf.setFont(
            "Helvetica-Bold",
            10
        )

        totals = [

            (
                "Subtotal",
                invoice.get(
                    "Subtotal",
                    0
                )
            ),

            (
                "Discount",
                invoice.get(
                    "Discount",
                    0
                )
            ),

            (
                "GST",
                invoice.get(
                    "GST",
                    0
                )
            ),

            (
                "Grand Total",
                invoice.get(
                    "Grand Total",
                    0
                )
            )
        ]

        for label, value in totals:

            pdf.drawString(
                350,
                y,
                label
            )

            pdf.drawString(
                470,
                y,
                f"Rs. {value}"
            )

            y -= 20

        y -= 20

        pdf.setFont(
            "Helvetica",
            8
        )

        pdf.drawString(
            50,
            y,
            "Goods once sold will not be taken back."
        )

        pdf.save()

        return path
